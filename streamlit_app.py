import streamlit as st
import pandas as pd
from collections import defaultdict, deque
from io import BytesIO
import hashlib
import urllib.parse

from streamlit_vis_network import streamlit_vis_network


# ============================================================
# Core Routing (Version A logic, refactored to use DataFrames)
# ============================================================

class CableNetwork:
    def __init__(self):
        self.graph = defaultdict(list)
        self.noise_levels = {}
        self.endpoints = defaultdict(list)
        self.endpoint_exposed = {}

    @staticmethod
    def _infer_noise_from_name(run_name: str):
        name = str(run_name or "")
        if "T6" in name:
            return {1}
        if "T4" in name:
            return {2}
        if "T3" in name:
            return {3}
        return set()

    @staticmethod
    def _parse_noise_levels(val, run_name=None):
        if pd.isna(val) or str(val).strip() == "":
            return CableNetwork._infer_noise_from_name(run_name)

        if isinstance(val, (int, float)) and not pd.isna(val):
            return {int(val)}

        s = str(val).strip()
        parts = [p.strip() for p in s.split(",")]
        out = set()
        for p in parts:
            if p:
                try:
                    out.add(int(float(p)))
                except ValueError:
                    pass
        return out

    def add_tray_or_conduit(self, name, noise_levels):
        levels = set(noise_levels) if isinstance(noise_levels, (set, list, tuple)) else {int(noise_levels)}
        self.noise_levels[name] = levels
        if name not in self.graph:
            self.graph[name] = []

    def connect(self, from_node, to_node):
        if from_node in self.noise_levels and to_node in self.noise_levels:
            self.graph[from_node].append((to_node, self.noise_levels[to_node]))
            self.graph[to_node].append((from_node, self.noise_levels[from_node]))

    def register_endpoint(self, device_name, tray_or_conduits, exposed=None):
        device_name = str(device_name).strip().lstrip("+")
        for item in str(tray_or_conduits).split(","):
            it = item.strip()
            if it:
                self.endpoints[device_name].append(it)
        if exposed is not None:
            self.endpoint_exposed[device_name] = bool(exposed)

    def find_route(self, starts, ends, allowed_noise_level):
        visited = set()
        queue = deque([[start] for start in starts])
        end_set = set(ends)

        while queue:
            path = queue.popleft()
            node = path[-1]
            if node in end_set:
                return path
            if node not in visited:
                visited.add(node)
                for neighbor, neighbor_levels in self.graph[node]:
                    if neighbor not in visited and allowed_noise_level in (neighbor_levels or set()):
                        queue.append(path + [neighbor])
        return None

    def build_from_dfs(self, trays_df: pd.DataFrame, connections_df: pd.DataFrame, endpoints_df: pd.DataFrame):
        for _, row in trays_df.iterrows():
            run_name = row.get("RunName", None)
            if pd.isna(run_name) or str(run_name).strip() == "":
                continue
            run_name = str(run_name).strip()
            levels = self._parse_noise_levels(row.get("Noise Level", None), run_name=run_name)
            if levels:
                self.add_tray_or_conduit(run_name, levels)

        for _, row in connections_df.iterrows():
            a = row.get("From", None)
            b = row.get("To", None)
            if pd.isna(a) or pd.isna(b):
                continue
            self.connect(str(a).strip(), str(b).strip())

        for _, row in endpoints_df.iterrows():
            device = row.get("device/panel", "")
            trays = row.get("tray/conduit(s)", "")
            exposed_flag = False
            try:
                val = row.iloc[2] if len(row) >= 3 else ""
            except Exception:
                val = ""
            if str(val).strip().lower() == "yes":
                exposed_flag = True
            self.register_endpoint(device, trays, exposed=exposed_flag)

    def route_cables_df(self, cables_df: pd.DataFrame) -> pd.DataFrame:
        cables = []
        seen = set()

        for _, row in cables_df.iterrows():
            sort = None
            if "Sort" in row and not pd.isna(row["Sort"]):
                try:
                    sort = int(row["Sort"])
                except Exception:
                    sort = None

            name = str(row.get("Cable number", ""))
            start = str(row.get("equipfrom", "")).strip().lstrip("+")
            end = str(row.get("equipto", "")).strip().lstrip("+")

            nl = row.get("Noise Level", None)
            if pd.isna(nl):
                continue
            try:
                noise_level = int(float(nl))
            except Exception:
                continue

            key = sort if sort is not None else name
            if key in seen:
                continue
            seen.add(key)

            cables.append({
                "sort": sort,
                "name": name,
                "start": start,
                "end": end,
                "noise_level": noise_level
            })

        results = []
        for cable in cables:
            nl = cable["noise_level"]

            start_trays = [t for t in self.endpoints.get(cable["start"], [])
                           if nl in self.noise_levels.get(t, set())]
            end_trays = [t for t in self.endpoints.get(cable["end"], [])
                         if nl in self.noise_levels.get(t, set())]

            if not start_trays and not end_trays:
                path_result = "Error: No endpoints with matching noise level (start & end)"
            elif not start_trays:
                path_result = "Error: No start endpoint with matching noise level"
            elif not end_trays:
                path_result = "Error: No end endpoint with matching noise level"
            else:
                path = self.find_route(start_trays, set(end_trays), nl)
                if path:
                    suffix = ""
                    if nl == 1:
                        suffix = "(T6)"
                    elif nl == 2:
                        suffix = "(T4)"

                    def format_node(node):
                        node = str(node)
                        if "LT" in node:
                            return f"{node}{suffix}"
                        if "CND" in node:
                            return node
                        return node

                    parts = [format_node(p) for p in path]

                    FROM_exposed = self.endpoint_exposed.get(cable["start"], False)
                    TO_exposed = self.endpoint_exposed.get(cable["end"], False)

                    if FROM_exposed:
                        parts.insert(0, "EXPOSED CONDUIT ROUTE")
                    if TO_exposed:
                        parts.append("EXPOSED CONDUIT ROUTE")

                    path_result = ",".join(parts)
                else:
                    path_result = "No valid route"

            results.append({
                "Sort": cable["sort"],
                "Cable number": cable["name"],
                "equipfrom": cable["start"],
                "equipto": cable["end"],
                "Noise Level": nl,
                "Via": path_result
            })

        return pd.DataFrame(results)


# ============================================================
# Excel I/O helpers
# ============================================================

REQUIRED_SHEETS = ["Tray", "Connections", "Endpoints", "Cables(input)"]

def load_excel_to_dfs(file_bytes: bytes) -> dict:
    xls = pd.ExcelFile(BytesIO(file_bytes))
    dfs = {}
    for sh in REQUIRED_SHEETS:
        if sh not in xls.sheet_names:
            raise ValueError(f"Missing required sheet: {sh}")
        dfs[sh] = pd.read_excel(xls, sh)
    return dfs

def write_updated_workbook_bytes(dfs: dict) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sh, df in dfs.items():
            df.to_excel(writer, sheet_name=sh, index=False)
    out.seek(0)
    return out.getvalue()

def write_routed_workbook_bytes(dfs: dict, routes_df: pd.DataFrame) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sh, df in dfs.items():
            df.to_excel(writer, sheet_name=sh, index=False)

        out_df = routes_df.copy()
        if "Cable number" in out_df.columns:
            out_df = out_df.rename(columns={"Cable number": "Cable Number"})
        out_df = out_df[["Sort", "Cable Number", "equipfrom", "equipto", "Noise Level", "Via"]]
        out_df = out_df.where(out_df.notna(), "")
        for c in out_df.columns:
            out_df[c] = out_df[c].astype(str)

        out_df.to_excel(writer, sheet_name="CableRoutes(output)", index=False)

        from openpyxl.styles import numbers
        wb = writer.book
        ws = wb["CableRoutes(output)"]
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        for col_name in ["Cable Number", "equipfrom", "equipto", "Via"]:
            if col_name in header_map:
                cidx = header_map[col_name]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    cell.value = "" if cell.value is None else str(cell.value)
                    cell.number_format = numbers.FORMAT_TEXT

    out.seek(0)
    return out.getvalue()

def validate_dfs(dfs: dict) -> list[str]:
    errors = []
    expected_cols = {
        "Tray": {"RunName"},
        "Connections": {"From", "To"},
        "Endpoints": {"device/panel", "tray/conduit(s)"},
        "Cables(input)": {"Cable number", "equipfrom", "equipto", "Noise Level"},
    }
    for sh, cols in expected_cols.items():
        missing = cols - set(dfs[sh].columns)
        if missing:
            errors.append(f"{sh}: missing columns {sorted(missing)}")
    return errors

def build_demo_workbook_bytes() -> bytes:
    tray = pd.DataFrame([
        {"RunName": "LT-01", "Noise Level": "1",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "LT-02", "Noise Level": "1",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "LT-03", "Noise Level": "1",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "CND-01","Noise Level": "1",   "X": pd.NA, "Y": pd.NA},

        {"RunName": "LT-11", "Noise Level": "2",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "LT-12", "Noise Level": "2",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "LT-13", "Noise Level": "2",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "CND-02","Noise Level": "2",   "X": pd.NA, "Y": pd.NA},

        {"RunName": "LT-20", "Noise Level": "1,2", "X": pd.NA, "Y": pd.NA},
        {"RunName": "CND-20","Noise Level": "1,2", "X": pd.NA, "Y": pd.NA},

        {"RunName": "LT-04", "Noise Level": "1",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "LT-14", "Noise Level": "2",   "X": pd.NA, "Y": pd.NA},
    ])

    connections = pd.DataFrame([
        {"From": "LT-01", "To": "LT-02"},
        {"From": "LT-02", "To": "LT-03"},
        {"From": "LT-02", "To": "LT-04"},
        {"From": "LT-03", "To": "CND-01"},

        {"From": "LT-11", "To": "LT-12"},
        {"From": "LT-12", "To": "LT-13"},
        {"From": "LT-12", "To": "LT-14"},
        {"From": "LT-13", "To": "CND-02"},

        {"From": "CND-01", "To": "LT-20"},
        {"From": "CND-02", "To": "LT-20"},
        {"From": "LT-20",  "To": "CND-20"},

        {"From": "LT-04",  "To": "LT-20"},
        {"From": "LT-14",  "To": "LT-20"},
    ])

    endpoints = pd.DataFrame([
        {"device/panel": "PANEL-A", "tray/conduit(s)": "LT-01,LT-11", "Exposed": ""},
        {"device/panel": "PANEL-B", "tray/conduit(s)": "LT-03,LT-13", "Exposed": "yes"},
        {"device/panel": "PANEL-C", "tray/conduit(s)": "LT-04",       "Exposed": ""},
        {"device/panel": "PANEL-D", "tray/conduit(s)": "LT-14",       "Exposed": ""},
    ])

    cables = pd.DataFrame([
        {"Sort": 1, "Cable number": "+CBL-0001", "equipfrom": "PANEL-A", "equipto": "PANEL-B", "Noise Level": 1},
        {"Sort": 2, "Cable number": "+CBL-0002", "equipfrom": "PANEL-A", "equipto": "PANEL-B", "Noise Level": 2},
        {"Sort": 3, "Cable number": "+CBL-0003", "equipfrom": "PANEL-C", "equipto": "PANEL-B", "Noise Level": 1},
        {"Sort": 4, "Cable number": "+CBL-0004", "equipfrom": "PANEL-D", "equipto": "PANEL-B", "Noise Level": 2},
    ])

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        tray.to_excel(writer, sheet_name="Tray", index=False)
        connections.to_excel(writer, sheet_name="Connections", index=False)
        endpoints.to_excel(writer, sheet_name="Endpoints", index=False)
        cables.to_excel(writer, sheet_name="Cables(input)", index=False)
    out.seek(0)
    return out.getvalue()


# ============================================================
# Endpoint lookup helpers
# ============================================================

def _normalize_token_list(s: str) -> list[str]:
    parts = [p.strip() for p in str(s or "").split(",")]
    return [p for p in parts if p]

def build_endpoint_to_trays_map(endpoints_df: pd.DataFrame) -> dict[str, list[str]]:
    m: dict[str, list[str]] = {}
    if endpoints_df is None or endpoints_df.empty:
        return m
    if "device/panel" not in endpoints_df.columns or "tray/conduit(s)" not in endpoints_df.columns:
        return m

    for _, r in endpoints_df.iterrows():
        ep = str(r.get("device/panel", "")).strip().lstrip("+")
        tc = r.get("tray/conduit(s)", "")
        if not ep:
            continue
        trays = _normalize_token_list(tc)
        if ep in m:
            m[ep].extend(trays)
        else:
            m[ep] = trays

    for k, v in list(m.items()):
        seen = set()
        out = []
        for x in v:
            if x not in seen:
                out.append(x)
                seen.add(x)
        m[k] = out
    return m

def build_tray_to_endpoints_map(endpoints_df: pd.DataFrame) -> dict[str, list[str]]:
    m: dict[str, list[str]] = {}
    if endpoints_df is None or endpoints_df.empty:
        return m
    if "device/panel" not in endpoints_df.columns or "tray/conduit(s)" not in endpoints_df.columns:
        return m

    for _, r in endpoints_df.iterrows():
        ep = str(r.get("device/panel", "")).strip().lstrip("+")
        tc = r.get("tray/conduit(s)", "")
        if not ep:
            continue
        trays = _normalize_token_list(tc)
        for t in trays:
            m.setdefault(t, []).append(ep)

    for k, v in list(m.items()):
        seen = set()
        out = []
        for x in v:
            if x not in seen:
                out.append(x)
                seen.add(x)
        m[k] = out
    return m


# ============================================================
# Route lookup helpers
# ============================================================

def compute_routes_df(tray_df: pd.DataFrame, connections_df: pd.DataFrame, endpoints_df: pd.DataFrame, cables_df: pd.DataFrame) -> pd.DataFrame:
    net = CableNetwork()
    net.build_from_dfs(tray_df, connections_df, endpoints_df)
    return net.route_cables_df(cables_df)

def _strip_route_suffix(node: str) -> str:
    s = str(node or "").strip()
    for suf in ["(T6)", "(T4)", "(T3)"]:
        if s.endswith(suf):
            s = s[: -len(suf)].strip()
    return s

def nodes_from_via_string(via: str) -> list[str]:
    if via is None or str(via).strip() == "":
        return []
    s = str(via).strip()
    if s.lower().startswith("error:"):
        return []
    if s.strip().lower() == "no valid route":
        return []
    parts = [p.strip() for p in s.split(",") if p.strip()]
    cleaned = []
    for p in parts:
        if p.upper() == "EXPOSED CONDUIT ROUTE":
            continue
        cleaned.append(_strip_route_suffix(p))
    return cleaned


# ============================================================
# Graph helpers (SVG nodes + seam fix + highlighting)
# ============================================================

ORANGE = "#FFA500"
GREEN  = "#00A651"
YELLOW = "#FFD200"
GRAY   = "#CFCFCF"

EDGE_COLOR = "#000000"
EDGE_WIDTH = 2

NODE_BORDER_COLOR = "#333333"
SVG_BORDER_WIDTH = 2

TRAY_SIDE = 70
TRAY_RADIUS = 14
IMAGE_SIZE = 32

PROBE_ID = "__POS_PROBE__"

# Very obvious highlight styling
HIGHLIGHT_BORDER_WIDTH = 10
HIGHLIGHT_BORDER_COLOR = "#FF00FF"
HIGHLIGHT_GLOW_SIZE = 22

def ensure_xy_columns(tray_df: pd.DataFrame) -> pd.DataFrame:
    tray_df = tray_df.copy()
    if "X" not in tray_df.columns:
        tray_df["X"] = pd.NA
    if "Y" not in tray_df.columns:
        tray_df["Y"] = pd.NA
    return tray_df

def noise_title(run_name: str, noise_val) -> str:
    lv = CableNetwork._parse_noise_levels(noise_val, run_name=run_name)
    if not lv:
        return "Noise Levels: N/A"
    return "Noise Levels: " + ",".join(str(x) for x in sorted(lv))

def infer_type_from_name(name: str) -> str:
    s = str(name).upper()
    if "CND" in s:
        return "Conduit"
    if "LT" in s:
        return "Tray"
    return "Node"

def svg_data_uri(svg: str) -> str:
    return "data:image/svg+xml;utf8," + urllib.parse.quote(svg)

def solid_rounded_square_svg(side: int, r: int, color: str, border: str = NODE_BORDER_COLOR) -> str:
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{side}" height="{side}" viewBox="0 0 {side} {side}">
      <rect x="0" y="0" width="{side}" height="{side}" rx="{r}" ry="{r}" fill="{color}"/>
      <rect x="1" y="1" width="{side-2}" height="{side-2}" rx="{r}" ry="{r}"
            fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def split_rounded_square_svg(side: int, r: int, left_color: str, right_color: str, border: str = NODE_BORDER_COLOR) -> str:
    half = side // 2
    left_w = half + 1
    right_x = half
    right_w = side - half
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{side}" height="{side}" viewBox="0 0 {side} {side}">
      <defs>
        <clipPath id="clipR">
          <rect x="0" y="0" width="{side}" height="{side}" rx="{r}" ry="{r}" />
        </clipPath>
      </defs>
      <g clip-path="url(#clipR)">
        <rect x="0" y="0" width="{left_w}" height="{side}" fill="{left_color}"/>
        <rect x="{right_x}" y="0" width="{right_w}" height="{side}" fill="{right_color}"/>
      </g>
      <rect x="1" y="1" width="{side-2}" height="{side-2}" rx="{r}" ry="{r}"
            fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def solid_circle_svg(d: int, color: str, border: str = NODE_BORDER_COLOR) -> str:
    r = d / 2
    rr = r - 1
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{d}" height="{d}" viewBox="0 0 {d} {d}">
      <circle cx="{r}" cy="{r}" r="{rr}" fill="{color}"/>
      <circle cx="{r}" cy="{r}" r="{rr}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def split_circle_svg(d: int, left_color: str, right_color: str, border: str = NODE_BORDER_COLOR) -> str:
    r = d / 2
    rr = r - 1
    left_w = int(r) + 1
    right_x = int(r)
    right_w = d - int(r)
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{d}" height="{d}" viewBox="0 0 {d} {d}">
      <defs>
        <clipPath id="clipC">
          <circle cx="{r}" cy="{r}" r="{rr}"/>
        </clipPath>
      </defs>
      <g clip-path="url(#clipC)">
        <rect x="0" y="0" width="{left_w}" height="{d}" fill="{left_color}"/>
        <rect x="{right_x}" y="0" width="{right_w}" height="{d}" fill="{right_color}"/>
      </g>
      <circle cx="{r}" cy="{r}" r="{rr}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def noise_color_kind(levels: set[int]) -> str:
    if levels == {1}:
        return "nl1"
    if levels == {2}:
        return "nl2"
    if levels == {3} or levels == {4}:
        return "nl34"
    if (1 in levels) and (2 in levels):
        return "mixed12"
    return "other"

def build_adjacency(connections_df: pd.DataFrame) -> dict[str, set[str]]:
    adj = defaultdict(set)
    if connections_df is None or connections_df.empty:
        return adj
    if "From" not in connections_df.columns or "To" not in connections_df.columns:
        return adj
    for _, r in connections_df.iterrows():
        a = r.get("From", None)
        b = r.get("To", None)
        if pd.isna(a) or pd.isna(b):
            continue
        a = str(a).strip()
        b = str(b).strip()
        if not a or not b:
            continue
        adj[a].add(b)
        adj[b].add(a)
    return adj

def neighborhood_nodes(adj: dict[str, set[str]], start: str, depth: int) -> set[str]:
    start = str(start).strip()
    if not start:
        return set()
    seen = {start}
    frontier = {start}
    for _ in range(max(0, int(depth))):
        nxt = set()
        for u in frontier:
            nxt |= set(adj.get(u, set()))
        nxt -= seen
        seen |= nxt
        frontier = nxt
        if not frontier:
            break
    return seen

def build_vis_nodes_edges(
    tray_df: pd.DataFrame,
    connections_df: pd.DataFrame,
    focus_node: str | None = None,
    focus_depth: int = 2,
    include_probe: bool = False,
    highlight_nodes: set[str] | None = None,
):
    tray_df = ensure_xy_columns(tray_df)
    highlight_nodes = set(highlight_nodes or [])

    focus_set = None
    if focus_node:
        adj = build_adjacency(connections_df)
        focus_set = neighborhood_nodes(adj, focus_node, focus_depth)

    nodes = []
    for _, r in tray_df.iterrows():
        rn = r.get("RunName", None)
        if pd.isna(rn) or str(rn).strip() == "":
            continue
        rn = str(rn).strip()

        if focus_set is not None and rn not in focus_set:
            continue

        is_focus = (focus_node is not None and rn == str(focus_node).strip())
        is_highlight = rn in highlight_nodes

        levels = CableNetwork._parse_noise_levels(r.get("Noise Level", None), run_name=rn)
        kind = noise_color_kind(levels)
        ntype = infer_type_from_name(rn)

        if ntype == "Tray":
            if kind == "nl1":
                svg = solid_rounded_square_svg(TRAY_SIDE, TRAY_RADIUS, ORANGE, NODE_BORDER_COLOR)
            elif kind == "nl2":
                svg = solid_rounded_square_svg(TRAY_SIDE, TRAY_RADIUS, GREEN, NODE_BORDER_COLOR)
            elif kind == "nl34":
                svg = solid_rounded_square_svg(TRAY_SIDE, TRAY_RADIUS, YELLOW, NODE_BORDER_COLOR)
            elif kind == "mixed12":
                svg = split_rounded_square_svg(TRAY_SIDE, TRAY_RADIUS, ORANGE, GREEN, NODE_BORDER_COLOR)
            else:
                svg = solid_rounded_square_svg(TRAY_SIDE, TRAY_RADIUS, GRAY, NODE_BORDER_COLOR)
        else:
            d = 80
            if kind == "nl1":
                svg = solid_circle_svg(d, ORANGE, NODE_BORDER_COLOR)
            elif kind == "nl2":
                svg = solid_circle_svg(d, GREEN, NODE_BORDER_COLOR)
            elif kind == "nl34":
                svg = solid_circle_svg(d, YELLOW, NODE_BORDER_COLOR)
            elif kind == "mixed12":
                svg = split_circle_svg(d, ORANGE, GREEN, NODE_BORDER_COLOR)
            else:
                svg = solid_circle_svg(d, GRAY, NODE_BORDER_COLOR)

        node = {
            "id": rn,
            "label": rn,
            "title": noise_title(rn, r.get("Noise Level", "")),
            "shape": "image",
            "image": svg_data_uri(svg),
            "size": IMAGE_SIZE,
            "borderWidth": 0,
            "font": {"vadjust": 0},
        }

        if is_highlight:
            node["borderWidth"] = HIGHLIGHT_BORDER_WIDTH
            node["color"] = {"border": HIGHLIGHT_BORDER_COLOR}
            node["shadow"] = {"enabled": True, "size": HIGHLIGHT_GLOW_SIZE, "x": 0, "y": 0}
            node["size"] = IMAGE_SIZE + 18
            node["font"] = {"vadjust": 0, "size": 20}

        if is_focus:
            node["borderWidth"] = max(node.get("borderWidth", 0), 3)
            node["color"] = {"border": "#000000"}

        x = r.get("X", pd.NA)
        y = r.get("Y", pd.NA)
        try:
            x = None if pd.isna(x) else float(x)
        except Exception:
            x = None
        try:
            y = None if pd.isna(y) else float(y)
        except Exception:
            y = None
        if x is not None and y is not None:
            node["x"] = x
            node["y"] = y

        nodes.append(node)

    node_ids = {n["id"] for n in nodes}

    edges = []
    seen = set()
    if connections_df is not None and not connections_df.empty and "From" in connections_df.columns and "To" in connections_df.columns:
        for _, r in connections_df.iterrows():
            a = r.get("From", None)
            b = r.get("To", None)
            if pd.isna(a) or pd.isna(b):
                continue
            a = str(a).strip()
            b = str(b).strip()
            if not a or not b:
                continue
            if a not in node_ids or b not in node_ids:
                continue
            key = tuple(sorted((a, b)))
            if key in seen:
                continue
            seen.add(key)
            edges.append({
                "id": f"{a}|||{b}",
                "from": a,
                "to": b,
                "title": f"{a} ↔ {b}",
                "color": EDGE_COLOR,
                "width": EDGE_WIDTH,
            })

    if include_probe and PROBE_ID not in node_ids:
        nodes.append({
            "id": PROBE_ID,
            "label": "",
            "title": "",
            "shape": "dot",
            "size": 1,
            "x": 0,
            "y": 0,
            "fixed": True,
            "physics": False,
            "color": {"background": "rgba(0,0,0,0)", "border": "rgba(0,0,0,0)"},
            "font": {"size": 0, "color": "rgba(0,0,0,0)"},
            "selected": True,
        })

    return nodes, edges

def apply_positions_to_tray(tray_df: pd.DataFrame, positions: dict) -> pd.DataFrame:
    tray_df = ensure_xy_columns(tray_df)
    if not positions:
        return tray_df

    tray_df = tray_df.copy()
    rn_series = tray_df["RunName"].astype(str).str.strip()

    for nid, xy in positions.items():
        if str(nid).strip() == PROBE_ID:
            continue

        x = y = None
        if isinstance(xy, dict):
            x = xy.get("x")
            y = xy.get("y")
        elif isinstance(xy, (list, tuple)) and len(xy) >= 2:
            x, y = xy[0], xy[1]

        try:
            x = float(x)
            y = float(y)
        except Exception:
            continue

        mask = rn_series == str(nid).strip()
        if mask.any():
            tray_df.loc[mask, "X"] = x
            tray_df.loc[mask, "Y"] = y

    return tray_df


# ============================================================
# DataFrame mutation helpers
# ============================================================

def df_delete_node(tray_df, connections_df, endpoints_df, node: str):
    node = str(node).strip()
    tray_df2 = tray_df.copy()
    tray_df2 = tray_df2[tray_df2["RunName"].astype(str).str.strip() != node].reset_index(drop=True)

    con2 = connections_df.copy()
    if "From" in con2.columns and "To" in con2.columns:
        con2 = con2[
            (con2["From"].astype(str).str.strip() != node) &
            (con2["To"].astype(str).str.strip() != node)
        ].reset_index(drop=True)

    ep2 = endpoints_df.copy()
    if "tray/conduit(s)" in ep2.columns:
        def remove_token(s):
            parts = [p.strip() for p in str(s).split(",") if p.strip()]
            parts = [p for p in parts if p != node]
            return ",".join(parts)
        ep2["tray/conduit(s)"] = ep2["tray/conduit(s)"].apply(remove_token)

    return tray_df2, con2, ep2

def df_delete_edge(connections_df, a: str, b: str):
    aa = str(a).strip()
    bb = str(b).strip()
    con = connections_df.copy()

    def is_match(r):
        x = str(r.get("From", "")).strip()
        y = str(r.get("To", "")).strip()
        return set([x, y]) == set([aa, bb])

    if ("From" in con.columns) and ("To" in con.columns):
        mask = con.apply(is_match, axis=1)
        con = con[~mask].reset_index(drop=True)
    return con

def df_add_edge(connections_df: pd.DataFrame, a: str, b: str) -> tuple[pd.DataFrame, bool]:
    a = str(a).strip()
    b = str(b).strip()
    if not a or not b or a == b:
        return connections_df, False

    con = connections_df.copy()

    if "From" not in con.columns or "To" not in con.columns:
        return connections_df, False

    def is_dup(r):
        x = str(r.get("From", "")).strip()
        y = str(r.get("To", "")).strip()
        return set([x, y]) == set([a, b])

    if not con.empty:
        try:
            if bool(con.apply(is_dup, axis=1).any()):
                return connections_df, False
        except Exception:
            pass

    new_row = {col: "" for col in con.columns}
    new_row["From"] = a
    new_row["To"] = b
    con = pd.concat([con, pd.DataFrame([new_row])], ignore_index=True)
    return con, True

def _compute_default_xy_near_network(tray_df: pd.DataFrame) -> tuple[float | None, float | None]:
    df = ensure_xy_columns(tray_df)
    xs = pd.to_numeric(df["X"], errors="coerce")
    ys = pd.to_numeric(df["Y"], errors="coerce")
    mask = xs.notna() & ys.notna()
    if mask.any():
        cx = float(xs[mask].mean())
        cy = float(ys[mask].mean())
        return cx + 140.0, cy
    return 200.0, 200.0

def df_add_node(tray_df: pd.DataFrame, name: str, noise_level_text: str = "", x=None, y=None) -> tuple[pd.DataFrame, bool]:
    name = str(name).strip()
    if not name:
        return tray_df, False

    df = ensure_xy_columns(tray_df.copy())

    if "RunName" not in df.columns:
        df["RunName"] = ""
    if "Noise Level" not in df.columns:
        df["Noise Level"] = ""

    if (df["RunName"].astype(str).str.strip() == name).any():
        return tray_df, False

    new_row = {col: "" for col in df.columns}
    new_row["RunName"] = name

    nl_text = str(noise_level_text or "").strip()
    if nl_text:
        new_row["Noise Level"] = nl_text

    if (x is None or str(x).strip() == "") and (y is None or str(y).strip() == ""):
        dx, dy = _compute_default_xy_near_network(df)
        new_row["X"] = dx
        new_row["Y"] = dy
    else:
        try:
            if x is not None and str(x).strip() != "":
                new_row["X"] = float(x)
        except Exception:
            pass
        try:
            if y is not None and str(y).strip() != "":
                new_row["Y"] = float(y)
        except Exception:
            pass

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    return df, True

def df_rename_node(tray_df, connections_df, endpoints_df, old: str, new: str):
    old = str(old).strip()
    new = str(new).strip()
    if not old or not new or old == new:
        return tray_df, connections_df, endpoints_df

    t = tray_df.copy()
    mask = t["RunName"].astype(str).str.strip() == old
    if mask.any():
        t.loc[mask, "RunName"] = new

    c = connections_df.copy()
    if "From" in c.columns:
        c.loc[c["From"].astype(str).str.strip() == old, "From"] = new
    if "To" in c.columns:
        c.loc[c["To"].astype(str).str.strip() == old, "To"] = new

    e = endpoints_df.copy()
    if "tray/conduit(s)" in e.columns:
        def rename_token(s):
            parts = [p.strip() for p in str(s).split(",") if p.strip()]
            parts = [new if p == old else p for p in parts]
            return ",".join(parts)
        e["tray/conduit(s)"] = e["tray/conduit(s)"].apply(rename_token)

    return t, c, e

def df_duplicate_node(tray_df, source_name: str, new_name: str):
    source_name = str(source_name).strip()
    new_name = str(new_name).strip()
    if not source_name or not new_name:
        return tray_df
    df = ensure_xy_columns(tray_df.copy())

    if (df["RunName"].astype(str).str.strip() == new_name).any():
        return df

    src_mask = df["RunName"].astype(str).str.strip() == source_name
    if not src_mask.any():
        return df

    src_row = df[src_mask].iloc[0].to_dict()
    src_row["RunName"] = new_name

    try:
        if not pd.isna(src_row.get("X")):
            src_row["X"] = float(src_row["X"]) + 50.0
        if not pd.isna(src_row.get("Y")):
            src_row["Y"] = float(src_row["Y"]) + 50.0
    except Exception:
        pass

    df = pd.concat([df, pd.DataFrame([src_row])], ignore_index=True)
    return df

def df_set_node_noise_level(tray_df: pd.DataFrame, node: str, noise_level_text: str) -> tuple[pd.DataFrame, bool]:
    node = str(node).strip()
    df = tray_df.copy()
    if "RunName" not in df.columns:
        return tray_df, False
    if "Noise Level" not in df.columns:
        df["Noise Level"] = ""

    mask = df["RunName"].astype(str).str.strip() == node
    if not mask.any():
        return tray_df, False

    df.loc[mask, "Noise Level"] = str(noise_level_text or "").strip()
    return df, True


# ============================================================
# Selection parsing helper (edge delete)
# ============================================================

def parse_selected_edge(sel_edge) -> tuple[str | None, str | None, str]:
    if sel_edge is None:
        return None, None, "None"

    if isinstance(sel_edge, dict):
        a = str(sel_edge.get("from", "")).strip()
        b = str(sel_edge.get("to", "")).strip()
        if a and b:
            return a, b, f"{a} ↔ {b}"
        sid = sel_edge.get("id")
        if isinstance(sid, str) and "|||" in sid:
            x, y = sid.split("|||", 1)
            return x.strip(), y.strip(), f"{x.strip()} ↔ {y.strip()}"
        return None, None, str(sel_edge)

    if isinstance(sel_edge, (list, tuple)):
        if len(sel_edge) >= 2:
            a = str(sel_edge[0]).strip()
            b = str(sel_edge[1]).strip()
            if a and b:
                return a, b, f"{a} ↔ {b}"
        return None, None, str(sel_edge)

    if isinstance(sel_edge, str):
        s = sel_edge.strip()
        if "|||" in s:
            x, y = s.split("|||", 1)
            return x.strip(), y.strip(), f"{x.strip()} ↔ {y.strip()}"
        return None, None, s

    return None, None, str(sel_edge)


# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="Cable Routing Webapp (Streamlit)", layout="wide")
st.title("Cable Routing Webapp (Streamlit)")

st.sidebar.header("Workbook")

st.session_state.setdefault("uploader_key_v", 0)
uploader_key = f"uploader_{st.session_state.uploader_key_v}"

source = st.sidebar.radio(
    "Choose workbook source",
    options=["Upload Excel (.xlsx)", "Use demo workbook"],
    index=0,
    key="workbook_source",
)

uploaded = None
demo_clicked = False

if source == "Upload Excel (.xlsx)":
    uploaded = st.sidebar.file_uploader("Upload Excel (.xlsx)", type=["xlsx"], key=uploader_key)
else:
    demo_clicked = st.sidebar.button("Load demo workbook", key="load_demo_btn", width="stretch")
    st.sidebar.caption("Loads a built-in example with multiple routes (Noise Levels 1 and 2 only).")
    demo_bytes_for_dl = build_demo_workbook_bytes()
    st.sidebar.download_button(
        "Download demo workbook (.xlsx)",
        data=demo_bytes_for_dl,
        file_name="demo_network_configuration.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.session_state.setdefault("tray_df", None)
st.session_state.setdefault("connections_df", None)
st.session_state.setdefault("endpoints_df", None)
st.session_state.setdefault("cables_df", None)
st.session_state.setdefault("routes_df", None)
st.session_state.setdefault("upload_hash", None)

st.session_state.setdefault("sel_nodes", [])
st.session_state.setdefault("sel_edges", [])

st.session_state.setdefault("focus_node", None)
st.session_state.setdefault("focus_depth", 2)

st.session_state.setdefault("graph_key_v", 0)

st.session_state.setdefault("layout_opt_active", False)
st.session_state.setdefault("layout_opt_last_positions", None)
st.session_state.setdefault("layout_opt_backup_xy", None)

st.session_state.setdefault("endpoint_highlight_nodes", set())
st.session_state.setdefault("endpoint_highlight_ep", None)

# NEW: route highlight state
st.session_state.setdefault("route_highlight_nodes", set())
st.session_state.setdefault("route_highlight_cable", None)

GRAPH_HEIGHT = 740

st.markdown(
    """
    <style>
      div[data-testid="stVerticalBlockBorderWrapper"]{
        border-radius: 10px !important;
      }
      .graphwrap, .graphwrap > div { width: 100% !important; }
      .graphwrap iframe { width: 100% !important; min-width: 100% !important; display: block !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

if st.sidebar.button("Clear workbook", key="clear_workbook_btn"):
    st.session_state.tray_df = None
    st.session_state.connections_df = None
    st.session_state.endpoints_df = None
    st.session_state.cables_df = None
    st.session_state.routes_df = None
    st.session_state.upload_hash = None
    st.session_state.sel_nodes = []
    st.session_state.sel_edges = []
    st.session_state.focus_node = None
    st.session_state.graph_key_v += 1
    st.session_state.uploader_key_v += 1
    st.session_state.layout_opt_active = False
    st.session_state.layout_opt_last_positions = None
    st.session_state.layout_opt_backup_xy = None
    st.session_state.endpoint_highlight_nodes = set()
    st.session_state.endpoint_highlight_ep = None
    st.session_state.route_highlight_nodes = set()
    st.session_state.route_highlight_cable = None
    st.rerun()

file_bytes = None
if uploaded is not None:
    file_bytes = uploaded.getvalue()
if demo_clicked:
    file_bytes = build_demo_workbook_bytes()

if file_bytes is not None:
    try:
        this_hash = hashlib.md5(file_bytes).hexdigest()
        if st.session_state.upload_hash != this_hash:
            loaded = load_excel_to_dfs(file_bytes)
            st.session_state.tray_df = ensure_xy_columns(loaded["Tray"])
            st.session_state.connections_df = loaded["Connections"]
            st.session_state.endpoints_df = loaded["Endpoints"]
            st.session_state.cables_df = loaded["Cables(input)"]
            st.session_state.routes_df = None
            st.session_state.upload_hash = this_hash
            st.session_state.sel_nodes = []
            st.session_state.sel_edges = []
            st.session_state.focus_node = None
            st.session_state.graph_key_v += 1
            st.session_state.layout_opt_active = False
            st.session_state.layout_opt_last_positions = None
            st.session_state.layout_opt_backup_xy = None
            st.session_state.endpoint_highlight_nodes = set()
            st.session_state.endpoint_highlight_ep = None
            st.session_state.route_highlight_nodes = set()
            st.session_state.route_highlight_cable = None
            st.sidebar.success("Workbook loaded.")
    except Exception as e:
        st.sidebar.error(f"Failed to load workbook: {e}")
        st.stop()

if st.session_state.tray_df is None:
    st.info("Upload an Excel workbook or load the demo workbook to begin.")
    st.stop()

st.session_state.tray_df = ensure_xy_columns(st.session_state.tray_df)

dfs = {
    "Tray": st.session_state.tray_df,
    "Connections": st.session_state.connections_df,
    "Endpoints": st.session_state.endpoints_df,
    "Cables(input)": st.session_state.cables_df,
}

val_errors = validate_dfs(dfs)
if val_errors:
    st.warning("Workbook validation warnings:")
    for e in val_errors:
        st.write(f"- {e}")

tab1, tab2, tab3, tab4, tabG, tab5 = st.tabs(
    ["Tray", "Connections", "Endpoints", "Cables(input)", "Graph Editor", "Routing Output"]
)

with tab1:
    st.subheader("Tray")
    st.caption("X and Y store node positions (saved manually from Graph Editor).")
    st.session_state.tray_df = st.data_editor(
        st.session_state.tray_df,
        num_rows="dynamic",
        width="stretch",
        key="tray_editor",
    )

with tab2:
    st.subheader("Connections")
    st.session_state.connections_df = st.data_editor(
        st.session_state.connections_df,
        num_rows="dynamic",
        width="stretch",
        key="connections_editor",
    )

with tab3:
    st.subheader("Endpoints")
    st.caption("Column C (3rd column) is interpreted as exposed yes/no (case-insensitive).")
    st.session_state.endpoints_df = st.data_editor(
        st.session_state.endpoints_df,
        num_rows="dynamic",
        width="stretch",
        key="endpoints_editor",
    )

with tab4:
    st.subheader("Cables(input)")
    st.session_state.cables_df = st.data_editor(
        st.session_state.cables_df,
        num_rows="dynamic",
        width="stretch",
        key="cables_editor",
    )

with tabG:
    st.subheader("Graph Editor")
    st.caption("Select a node/edge to edit it on the right. Use Search to focus/highlight a node.")

    graph_col, tools_col = st.columns([3, 1], gap="large")

    node_names = (
        st.session_state.tray_df["RunName"]
        .dropna()
        .astype(str)
        .map(str.strip)
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )
    node_names = sorted(node_names, key=lambda s: s.lower())

    endpoint_to_trays = build_endpoint_to_trays_map(st.session_state.endpoints_df)
    tray_to_endpoints = build_tray_to_endpoints_map(st.session_state.endpoints_df)
    endpoint_names = sorted(endpoint_to_trays.keys(), key=lambda s: s.lower())

    selection = None

    with graph_col:
        graph_box = st.container(border=True)
        with graph_box:
            st.markdown('<div class="graphwrap">', unsafe_allow_html=True)

            combined_highlight = set(st.session_state.endpoint_highlight_nodes or set()) | set(st.session_state.route_highlight_nodes or set())

            nodes, edges = build_vis_nodes_edges(
                st.session_state.tray_df,
                st.session_state.connections_df,
                focus_node=st.session_state.focus_node,
                focus_depth=int(st.session_state.focus_depth),
                include_probe=bool(st.session_state.layout_opt_active),
                highlight_nodes=combined_highlight,
            )

            if st.session_state.layout_opt_active:
                options = {
                    "physics": {
                        "enabled": True,
                        "stabilization": {
                            "enabled": True,
                            "iterations": 1600,
                            "updateInterval": 50,
                            "fit": True,
                        },
                        "solver": "forceAtlas2Based",
                        "forceAtlas2Based": {
                            "gravitationalConstant": -260,
                            "centralGravity": 0.01,
                            "springLength": 240,
                            "springConstant": 0.02,
                            "damping": 0.4,
                            "avoidOverlap": 1.0,
                        },
                        "minVelocity": 0.20,
                        "maxVelocity": 70,
                        "timestep": 0.5,
                    },
                    "layout": {"improvedLayout": True},
                    "interaction": {
                        "dragNodes": True,
                        "dragView": True,
                        "zoomView": True,
                        "hover": True,
                        "multiselect": False,
                        "selectConnectedEdges": True,
                    },
                    "nodes": {"font": {"vadjust": 0}},
                    "edges": {"smooth": False, "color": EDGE_COLOR, "width": EDGE_WIDTH},
                }
            else:
                options = {
                    "physics": {"enabled": False},
                    "layout": {"improvedLayout": True},
                    "interaction": {
                        "dragNodes": True,
                        "dragView": True,
                        "zoomView": True,
                        "hover": True,
                        "multiselect": False,
                        "selectConnectedEdges": True,
                    },
                    "nodes": {"font": {"vadjust": 0}},
                    "edges": {"smooth": False, "color": EDGE_COLOR, "width": EDGE_WIDTH},
                }

            selection = streamlit_vis_network(
                nodes,
                edges,
                height=GRAPH_HEIGHT,
                options=options,
                key=f"vis_network_graph_{st.session_state.graph_key_v}",
            )

            st.markdown("</div>", unsafe_allow_html=True)

        if selection:
            try:
                sel_nodes, sel_edges, _pos = selection
            except Exception:
                sel_nodes, sel_edges, _pos = [], [], None

            st.session_state.layout_opt_last_positions = _pos if isinstance(_pos, dict) else st.session_state.layout_opt_last_positions

            cleaned_nodes = [n for n in (sel_nodes or []) if str(n).strip() != PROBE_ID]
            st.session_state.sel_nodes = cleaned_nodes
            st.session_state.sel_edges = sel_edges or []
        else:
            st.session_state.sel_nodes = []
            st.session_state.sel_edges = []

    with tools_col:
        tools_box = st.container(height=GRAPH_HEIGHT, border=True)
        with tools_box:
            st.markdown("### Selection")

            sel_nodes = st.session_state.sel_nodes or []
            sel_edges = st.session_state.sel_edges or []

            if sel_nodes:
                node_id = str(sel_nodes[0]).strip()
                tdf = st.session_state.tray_df
                row = tdf[tdf["RunName"].astype(str).str.strip() == node_id]
                noise_val = ""
                x_val = y_val = None
                if not row.empty:
                    noise_val = row.iloc[0].get("Noise Level", "")
                    x_val = row.iloc[0].get("X", pd.NA)
                    y_val = row.iloc[0].get("Y", pd.NA)

                st.markdown("**Selected Node**")
                st.write(f"**Name:** {node_id}")
                st.write(f"**Type:** {infer_type_from_name(node_id)}")
                st.write(f"**Noise Level:** {noise_val if str(noise_val).strip() else 'N/A'}")
                st.write(f"**X:** {'' if pd.isna(x_val) else x_val}")
                st.write(f"**Y:** {'' if pd.isna(y_val) else y_val}")

                eps = tray_to_endpoints.get(node_id, [])
                if eps:
                    st.caption("Endpoints connected to this tray/conduit (from Endpoints sheet):")
                    st.write(", ".join(eps))
                else:
                    st.caption("No endpoints reference this tray/conduit (from Endpoints sheet).")

                st.divider()

                st.markdown("**Connect this node**")
                connect_options = [""] + [n for n in node_names if n != node_id]
                target = st.selectbox(
                    "Connect to (type to search)",
                    options=connect_options,
                    index=0,
                    key="sel_connect_target",
                )

                c_add, c_del = st.columns(2)
                with c_add:
                    if st.button("Add connection", key="sel_add_conn_btn", width="stretch"):
                        tgt = (target or "").strip()
                        if not tgt:
                            st.warning("Choose a target node.")
                        else:
                            new_con, added = df_add_edge(st.session_state.connections_df, node_id, tgt)
                            if added:
                                st.session_state.connections_df = new_con
                                st.session_state.routes_df = None
                                st.success(f"Added connection: {node_id} ↔ {tgt}")
                                st.session_state.graph_key_v += 1
                                st.rerun()
                            else:
                                st.info("That connection already exists (or could not be added).")
                with c_del:
                    if st.button("Delete connection", key="sel_del_conn_btn", width="stretch"):
                        tgt = (target or "").strip()
                        if not tgt:
                            st.warning("Choose a target node.")
                        else:
                            before = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                            st.session_state.connections_df = df_delete_edge(st.session_state.connections_df, node_id, tgt)
                            after = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                            st.session_state.routes_df = None
                            if before == after:
                                st.info("No matching connection was found to delete.")
                            else:
                                st.success(f"Deleted connection: {node_id} ↔ {tgt}")
                                st.session_state.graph_key_v += 1
                            st.rerun()

                st.divider()

                if st.button("Delete node", key="sel_delete_node", width="stretch"):
                    t, c, e = df_delete_node(
                        st.session_state.tray_df,
                        st.session_state.connections_df,
                        st.session_state.endpoints_df,
                        node_id
                    )
                    st.session_state.tray_df = ensure_xy_columns(t)
                    st.session_state.connections_df = c
                    st.session_state.endpoints_df = e
                    st.session_state.routes_df = None
                    if st.session_state.focus_node == node_id:
                        st.session_state.focus_node = None
                    st.session_state.sel_nodes = []
                    st.session_state.sel_edges = []
                    st.success("Node deleted.")
                    st.session_state.graph_key_v += 1
                    st.rerun()

                st.markdown("**Rename**")
                new_name = st.text_input("New name", value=node_id, key="sel_rename_node_new")
                if st.button("Rename node", key="sel_rename_node_btn", width="stretch"):
                    if new_name.strip() and new_name.strip() != node_id:
                        if (st.session_state.tray_df["RunName"].astype(str).str.strip() == new_name.strip()).any():
                            st.error("That name already exists.")
                        else:
                            t, c, e = df_rename_node(
                                st.session_state.tray_df,
                                st.session_state.connections_df,
                                st.session_state.endpoints_df,
                                node_id,
                                new_name.strip()
                            )
                            st.session_state.tray_df = ensure_xy_columns(t)
                            st.session_state.connections_df = c
                            st.session_state.endpoints_df = e
                            st.session_state.routes_df = None
                            if st.session_state.focus_node == node_id:
                                st.session_state.focus_node = new_name.strip()
                            st.session_state.sel_nodes = [new_name.strip()]
                            st.session_state.sel_edges = []
                            st.success("Node renamed.")
                            st.session_state.graph_key_v += 1
                            st.rerun()
                    else:
                        st.warning("Enter a different non-empty name.")

                st.markdown("**Duplicate**")
                dup_name = st.text_input("Duplicate as", value=f"{node_id}_COPY", key="sel_dup_node_new")
                if st.button("Duplicate node", key="sel_dup_node_btn", width="stretch"):
                    if dup_name.strip():
                        if (st.session_state.tray_df["RunName"].astype(str).str.strip() == dup_name.strip()).any():
                            st.error("That duplicate name already exists.")
                        else:
                            st.session_state.tray_df = df_duplicate_node(st.session_state.tray_df, node_id, dup_name.strip())
                            st.session_state.routes_df = None
                            st.success("Node duplicated.")
                            st.session_state.graph_key_v += 1
                            st.rerun()
                    else:
                        st.warning("Enter a name for the duplicate.")

                st.divider()
                st.markdown("**Edit noise level**")

                preset_map = {
                    "": "(leave as-is)",
                    "1": "1",
                    "2": "2",
                    "1,2": "1,2",
                    "2,1": "2,1",
                }
                presets = list(preset_map.keys())

                current_text = str(noise_val or "").strip()
                preset_index = 0
                if current_text in presets:
                    preset_index = presets.index(current_text)

                nl_preset_sel = st.selectbox(
                    "Preset",
                    options=presets,
                    index=preset_index,
                    key="sel_noise_preset",
                    format_func=lambda x: preset_map.get(x, x),
                )

                nl_custom_sel = st.text_input(
                    "Custom (optional: overrides preset if non-empty)",
                    value="",
                    key="sel_noise_custom",
                    placeholder="e.g. 1 or 2 or 1,2",
                )

                if st.button("Apply noise level", key="apply_noise_btn", width="stretch"):
                    new_text = (nl_custom_sel or "").strip() if (nl_custom_sel or "").strip() else str(nl_preset_sel or "").strip()
                    if new_text == "(leave as-is)":
                        new_text = current_text

                    new_df, ok = df_set_node_noise_level(st.session_state.tray_df, node_id, new_text)
                    if ok:
                        st.session_state.tray_df = ensure_xy_columns(new_df)
                        st.session_state.routes_df = None
                        st.success(f"Updated noise level for {node_id} to: {new_text if new_text else '(blank)'}")
                        st.session_state.graph_key_v += 1
                        st.rerun()
                    else:
                        st.error("Could not update noise level for the selected node.")

            elif sel_edges:
                raw_edge = sel_edges[0]
                a, b, disp = parse_selected_edge(raw_edge)

                st.markdown("**Selected Connection**")
                st.write(f"**Connection:** {disp}")

                if not (a and b):
                    st.caption("Could not parse endpoints for this connection (unexpected format).")
                    st.code(str(raw_edge))
                else:
                    st.write(f"**From:** {a}")
                    st.write(f"**To:** {b}")
                    st.divider()

                    if st.button("Delete connection", key="sel_delete_edge", width="stretch"):
                        before = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                        st.session_state.connections_df = df_delete_edge(st.session_state.connections_df, a, b)
                        after = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                        st.session_state.routes_df = None

                        st.session_state.sel_nodes = []
                        st.session_state.sel_edges = []

                        if before == after:
                            st.info("No matching connection was found to delete (already removed?).")
                        else:
                            st.success("Connection deleted.")
                        st.session_state.graph_key_v += 1
                        st.rerun()

                    st.caption("This deletes the row from the Connections sheet (undirected match).")

            else:
                st.info("Select a node or connection in the graph to edit it here.")

            st.divider()

            st.markdown("### Search / Focus")

            st.session_state.focus_depth = st.slider(
                "Neighborhood depth",
                min_value=1,
                max_value=6,
                value=int(st.session_state.focus_depth),
                step=1,
            )

            focus_pick = st.selectbox(
                "Find a node (type to search)",
                options=[""] + node_names,
                index=0 if not st.session_state.focus_node else (
                    node_names.index(st.session_state.focus_node) + 1
                    if st.session_state.focus_node in node_names else 0
                ),
                key="focus_pick_select",
            )

            c1, c2 = st.columns(2)
            with c1:
                if st.button("Focus", width="stretch", key="focus_btn"):
                    if focus_pick and focus_pick.strip():
                        st.session_state.focus_node = focus_pick.strip()
                        st.session_state.sel_nodes = [st.session_state.focus_node]
                        st.session_state.sel_edges = []
                        st.session_state.graph_key_v += 1
                        st.rerun()
                    else:
                        st.warning("Pick a node to focus.")
            with c2:
                if st.button("Clear", width="stretch", key="clear_focus_btn"):
                    st.session_state.focus_node = None
                    st.session_state.graph_key_v += 1
                    st.rerun()

            if st.session_state.focus_node:
                st.caption(f"Focused on: **{st.session_state.focus_node}** (showing local neighborhood)")

            st.divider()

            st.markdown("### Connection")

            conn_from = st.selectbox("From (type to search)", options=[""] + node_names, index=0, key="conn_from")
            conn_to = st.selectbox("To (type to search)", options=[""] + node_names, index=0, key="conn_to")

            b_add, b_del = st.columns(2)
            with b_add:
                if st.button("Add", width="stretch", key="add_edge_btn"):
                    a = (conn_from or "").strip()
                    b = (conn_to or "").strip()
                    if not a or not b:
                        st.warning("Choose both From and To.")
                    elif a == b:
                        st.warning("From and To must be different.")
                    else:
                        new_con, added = df_add_edge(st.session_state.connections_df, a, b)
                        if added:
                            st.session_state.connections_df = new_con
                            st.session_state.routes_df = None
                            st.success(f"Added connection: {a} ↔ {b}")
                            st.session_state.graph_key_v += 1
                            st.rerun()
                        else:
                            st.info("That connection already exists (or could not be added).")
            with b_del:
                if st.button("Delete", width="stretch", key="delete_edge_btn"):
                    a = (conn_from or "").strip()
                    b = (conn_to or "").strip()
                    if not a or not b:
                        st.warning("Choose both From and To.")
                    elif a == b:
                        st.warning("From and To must be different.")
                    else:
                        before = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                        st.session_state.connections_df = df_delete_edge(st.session_state.connections_df, a, b)
                        after = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                        st.session_state.routes_df = None
                        if before == after:
                            st.info("No matching connection was found to delete.")
                        else:
                            st.session_state.sel_edges = []
                            st.session_state.sel_nodes = []
                            st.success(f"Deleted connection: {a} ↔ {b}")
                            st.session_state.graph_key_v += 1
                        st.rerun()

            st.divider()

            st.markdown("### Add node")

            new_node_name = st.text_input("New node name", value="", key="add_node_name")
            nl_preset = st.selectbox("Noise level (preset)", options=["(leave blank)", "1", "2", "1,2"], index=0, key="add_node_nl_preset")
            nl_custom = st.text_input("Noise level (custom text, optional)", value="", key="add_node_nl_custom")

            p1, p2 = st.columns(2)
            with p1:
                new_x = st.text_input("X (optional)", value="", key="add_node_x")
            with p2:
                new_y = st.text_input("Y (optional)", value="", key="add_node_y")

            if st.button("Add node", width="stretch", key="add_node_btn"):
                name = (new_node_name or "").strip()
                if not name:
                    st.warning("Enter a node name.")
                else:
                    if (nl_custom or "").strip():
                        nl_text = nl_custom.strip()
                    else:
                        nl_text = "" if nl_preset == "(leave blank)" else nl_preset

                    new_df, added = df_add_node(
                        st.session_state.tray_df,
                        name=name,
                        noise_level_text=nl_text,
                        x=new_x,
                        y=new_y,
                    )

                    if not added:
                        st.error("Could not add node (it may already exist).")
                    else:
                        st.session_state.tray_df = ensure_xy_columns(new_df)
                        st.session_state.routes_df = None
                        st.session_state.focus_node = None
                        st.session_state.sel_nodes = [name]
                        st.session_state.sel_edges = []
                        st.success(f"Added node: {name}")
                        st.session_state.graph_key_v += 1
                        st.rerun()

            st.divider()

            st.markdown("### Delete node (autocomplete)")

            del_pick = st.selectbox("Node to delete (type to search)", options=[""] + node_names, index=0, key="delete_node_pick")
            if st.button("Delete selected node", width="stretch", key="delete_node_autofill_btn"):
                if not del_pick.strip():
                    st.warning("Choose a node to delete.")
                else:
                    node_id = del_pick.strip()
                    t, c, e = df_delete_node(
                        st.session_state.tray_df,
                        st.session_state.connections_df,
                        st.session_state.endpoints_df,
                        node_id
                    )
                    st.session_state.tray_df = ensure_xy_columns(t)
                    st.session_state.connections_df = c
                    st.session_state.endpoints_df = e
                    st.session_state.routes_df = None
                    if st.session_state.focus_node == node_id:
                        st.session_state.focus_node = None
                    if st.session_state.sel_nodes and st.session_state.sel_nodes[0] == node_id:
                        st.session_state.sel_nodes = []
                        st.session_state.sel_edges = []
                    st.success(f"Deleted node: {node_id}")
                    st.session_state.graph_key_v += 1
                    st.rerun()

            st.divider()

            # ====================================================
            # Endpoint lookup (second-last, above Route lookup + Layout)
            # ====================================================
            st.markdown("### Endpoint lookup")

            ep_pick = st.selectbox(
                "Choose an endpoint/device (type to search)",
                options=[""] + endpoint_names,
                index=0,
                key="endpoint_lookup_pick",
            )

            e1, e2 = st.columns([1, 1])
            with e1:
                if st.button("Highlight connected trays/conduits", width="stretch", key="endpoint_lookup_highlight_btn"):
                    if not (ep_pick or "").strip():
                        st.warning("Pick an endpoint/device.")
                    else:
                        ep = (ep_pick or "").strip().lstrip("+")
                        trays = endpoint_to_trays.get(ep, [])
                        trays_present = [t for t in trays if t in node_names]
                        st.session_state.endpoint_highlight_nodes = set(trays_present)
                        st.session_state.endpoint_highlight_ep = ep
                        st.session_state.graph_key_v += 1
                        if trays_present:
                            st.success(f"Highlighted {len(trays_present)} tray/conduit node(s) for {ep}.")
                        else:
                            st.info("No trays/conduits for that endpoint are present in Tray.RunName (nothing to highlight).")
                        st.rerun()

            with e2:
                if st.button("Clear endpoint highlight", width="stretch", key="endpoint_lookup_clear_highlight_btn"):
                    st.session_state.endpoint_highlight_nodes = set()
                    st.session_state.endpoint_highlight_ep = None
                    st.session_state.graph_key_v += 1
                    st.success("Cleared endpoint highlight.")
                    st.rerun()

            if st.session_state.endpoint_highlight_ep:
                st.caption(f"Endpoint highlight active for: **{st.session_state.endpoint_highlight_ep}**")
                if st.session_state.endpoint_highlight_nodes:
                    st.write(", ".join(sorted(st.session_state.endpoint_highlight_nodes)))
                else:
                    st.write("No nodes are currently highlighted for this endpoint.")

            st.divider()

            # ====================================================
            # NEW: Route lookup (immediately underneath endpoint lookup)
            # ====================================================
            st.markdown("### Route lookup")

            routes_df = st.session_state.routes_df
            if routes_df is None:
                st.info("Routes are not computed yet. Compute them to enable route lookup/highlighting.")
                if st.button("Compute routes now", width="stretch", key="compute_routes_in_graph_editor_btn"):
                    try:
                        st.session_state.routes_df = compute_routes_df(
                            st.session_state.tray_df,
                            st.session_state.connections_df,
                            st.session_state.endpoints_df,
                            st.session_state.cables_df,
                        )
                        st.success("Routes computed. You can now use Route lookup.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Routing failed: {e}")
            else:
                # Build cable list for selectbox
                try:
                    ccol = "Cable number" if "Cable number" in routes_df.columns else ("Cable Number" if "Cable Number" in routes_df.columns else None)
                    if ccol is None:
                        cable_options = []
                    else:
                        cable_options = (
                            routes_df[ccol]
                            .astype(str)
                            .fillna("")
                            .map(str.strip)
                            .replace("", pd.NA)
                            .dropna()
                            .unique()
                            .tolist()
                        )
                        cable_options = sorted(cable_options, key=lambda s: s.lower())
                except Exception:
                    cable_options = []

                cable_pick = st.selectbox(
                    "Choose a routed cable (type to search)",
                    options=[""] + cable_options,
                    index=0,
                    key="route_lookup_pick",
                )

                r1, r2 = st.columns([1, 1])
                with r1:
                    if st.button("Highlight route path", width="stretch", key="route_lookup_highlight_btn"):
                        if not (cable_pick or "").strip():
                            st.warning("Pick a cable.")
                        else:
                            cnum = (cable_pick or "").strip()
                            st.session_state.route_highlight_cable = cnum

                            # find row
                            df = routes_df.copy()
                            col = "Cable number" if "Cable number" in df.columns else ("Cable Number" if "Cable Number" in df.columns else None)
                            if col is None or "Via" not in df.columns:
                                st.error("Routes table is missing required columns (Cable number/Cable Number, Via).")
                            else:
                                mask = df[col].astype(str).map(str.strip) == cnum
                                if not mask.any():
                                    st.warning("Could not find that cable in the routed output.")
                                else:
                                    via = df.loc[mask, "Via"].iloc[0]
                                    path_nodes = nodes_from_via_string(via)

                                    # only highlight nodes that exist in the graph
                                    present = [n for n in path_nodes if n in node_names]
                                    st.session_state.route_highlight_nodes = set(present)

                                    st.session_state.graph_key_v += 1
                                    if present:
                                        st.success(f"Highlighted {len(present)} node(s) along the route for {cnum}.")
                                    else:
                                        st.info("Route did not contain any Tray.RunName nodes to highlight (or it was an error/no-route).")
                                    st.rerun()

                with r2:
                    if st.button("Clear route highlight", width="stretch", key="route_lookup_clear_btn"):
                        st.session_state.route_highlight_nodes = set()
                        st.session_state.route_highlight_cable = None
                        st.session_state.graph_key_v += 1
                        st.success("Cleared route highlight.")
                        st.rerun()

                if st.session_state.route_highlight_cable:
                    st.caption(f"Route highlight active for: **{st.session_state.route_highlight_cable}**")
                    if st.session_state.route_highlight_nodes:
                        st.write(", ".join(sorted(st.session_state.route_highlight_nodes)))
                    else:
                        st.write("No nodes are currently highlighted for this route.")

            st.divider()

            st.markdown("### Layout")

            if st.button("Home / Recenter graph view", key="home_recenter_btn", width="stretch"):
                st.session_state.focus_node = None
                st.session_state.sel_nodes = []
                st.session_state.sel_edges = []
                st.session_state.graph_key_v += 1
                st.success("Recentered graph view.")
                st.rerun()

            if not st.session_state.layout_opt_active:
                if st.button("Optimize layout (turns on physics)", key="opt_turn_on_physics_btn", width="stretch"):
                    tdf = ensure_xy_columns(st.session_state.tray_df).copy()
                    backup = {
                        "RunName": tdf["RunName"].astype(str).tolist(),
                        "X": tdf["X"].tolist(),
                        "Y": tdf["Y"].tolist(),
                    }
                    st.session_state.layout_opt_backup_xy = backup

                    st.session_state.tray_df["X"] = pd.NA
                    st.session_state.tray_df["Y"] = pd.NA

                    st.session_state.routes_df = None
                    st.session_state.focus_node = None
                    st.session_state.sel_nodes = []
                    st.session_state.sel_edges = []

                    st.session_state.layout_opt_active = True
                    st.session_state.layout_opt_last_positions = None
                    st.session_state.graph_key_v += 1
                    st.success("Physics enabled for optimization. Let it settle, then save or cancel.")
                    st.rerun()
            else:
                st.info(
                    "Optimization is active (physics ON).\n\n"
                    "✅ Let the graph settle, then click **Save optimized positions (turns off physics)**.\n"
                    "↩️ Or click **Cancel optimization (turns off physics)** to revert to the previous layout."
                )

                if st.button("Save optimized positions (turns off physics)", key="save_opt_positions_btn", width="stretch"):
                    positions = st.session_state.layout_opt_last_positions

                    if positions and isinstance(positions, dict) and len(positions) > 0:
                        st.session_state.tray_df = apply_positions_to_tray(st.session_state.tray_df, positions)
                        st.session_state.layout_opt_active = False
                        st.session_state.layout_opt_last_positions = None
                        st.session_state.layout_opt_backup_xy = None
                        st.session_state.graph_key_v += 1
                        st.success(f"Saved optimized positions for {len(positions)} node(s). Physics is now OFF.")
                        st.rerun()
                    else:
                        st.warning(
                            "No positions were returned yet.\n\n"
                            "✅ Wait a moment for stabilization, then click **Save optimized positions** again."
                        )

                if st.button("Cancel optimization (turns off physics)", key="cancel_opt_btn", width="stretch"):
                    backup = st.session_state.layout_opt_backup_xy
                    if backup and "RunName" in backup:
                        tdf = ensure_xy_columns(st.session_state.tray_df).copy()
                        rn = tdf["RunName"].astype(str).tolist()
                        bmap = {str(n).strip(): (backup["X"][i], backup["Y"][i]) for i, n in enumerate(backup["RunName"])}

                        xs = []
                        ys = []
                        for n in rn:
                            x, y = bmap.get(str(n).strip(), (pd.NA, pd.NA))
                            xs.append(x)
                            ys.append(y)
                        tdf["X"] = xs
                        tdf["Y"] = ys
                        st.session_state.tray_df = tdf

                    st.session_state.layout_opt_active = False
                    st.session_state.layout_opt_last_positions = None
                    st.session_state.layout_opt_backup_xy = None
                    st.session_state.graph_key_v += 1
                    st.success("Optimization canceled. Previous layout restored. Physics is now OFF.")
                    st.rerun()

            if st.button("Save current node positions to Tray (X/Y)", key="save_positions_btn", width="stretch"):
                positions = None
                if selection:
                    try:
                        _sn, _se, positions = selection
                    except Exception:
                        positions = None

                if positions and isinstance(positions, dict) and len(positions) > 0:
                    st.session_state.tray_df = apply_positions_to_tray(st.session_state.tray_df, positions)
                    st.success(f"Saved positions for {len(positions)} node(s). You can keep dragging and save again anytime.")
                    st.rerun()
                else:
                    st.warning(
                        "No positions were returned yet.\n\n"
                        "✅ Drag a node, then click an empty area of the graph once, then click **Save** again."
                    )

            if st.button("Clear saved positions (blank X/Y)", key="clear_positions_btn", width="stretch"):
                st.session_state.tray_df["X"] = pd.NA
                st.session_state.tray_df["Y"] = pd.NA
                st.session_state.routes_df = None
                st.session_state.layout_opt_active = False
                st.session_state.layout_opt_last_positions = None
                st.session_state.layout_opt_backup_xy = None
                st.session_state.graph_key_v += 1
                st.success("Cleared Tray.X and Tray.Y.")
                st.rerun()

with tab5:
    st.subheader("Route cables")
    colA, colB, colC = st.columns([1, 1, 2])

    with colA:
        if st.button("Validate", key="validate_btn"):
            dfs_for_validate = {
                "Tray": st.session_state.tray_df,
                "Connections": st.session_state.connections_df,
                "Endpoints": st.session_state.endpoints_df,
                "Cables(input)": st.session_state.cables_df,
            }
            errs = validate_dfs(dfs_for_validate)
            if errs:
                st.error("Validation issues found:")
                for e in errs:
                    st.write(f"- {e}")
            else:
                st.success("Basic validation passed.")

    with colB:
        if st.button("Route Now", key="route_btn"):
            try:
                st.session_state.routes_df = compute_routes_df(
                    st.session_state.tray_df,
                    st.session_state.connections_df,
                    st.session_state.endpoints_df,
                    st.session_state.cables_df,
                )
                st.success("Routing complete.")
            except Exception as e:
                st.error(f"Routing failed: {e}")

    routes_df = st.session_state.routes_df
    if routes_df is not None:
        st.dataframe(routes_df, width="stretch")

        dfs_for_export = {
            "Tray": st.session_state.tray_df,
            "Connections": st.session_state.connections_df,
            "Endpoints": st.session_state.endpoints_df,
            "Cables(input)": st.session_state.cables_df,
        }

        updated_bytes = write_updated_workbook_bytes(dfs_for_export)
        routed_bytes = write_routed_workbook_bytes(dfs_for_export, routes_df)

        st.download_button(
            "Download UPDATED workbook (edited sheets only)",
            data=updated_bytes,
            file_name="network_configuration_updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.download_button(
            "Download ROUTED workbook (includes CableRoutes(output))",
            data=routed_bytes,
            file_name="network_configuration_routed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("Click **Route Now** to generate CableRoutes(output).")
